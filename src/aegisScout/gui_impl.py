"""
aegisScout GUI — PyWebView bridge.

SECURITY NOTE:
  The Python API exposed to JS (`GuiApi`) NEVER returns API keys, passwords,
  or tokens. Only booleans and safe non-sensitive configuration are surfaced
  via `is_configured()` (and the backward-compat `get_settings()`).

  Sensitive writes (API keys, passwords) must be performed by the user
  re-typing the value into the GUI form; the Python side accepts and persists
  them but never echoes them back to the JS layer.
"""

from __future__ import annotations

import asyncio
import os
import webview
import json
import re
import sys
import ctypes
import time
from pathlib import Path
import threading
from datetime import datetime, timezone, timedelta
from sqlmodel import Session, select, text
from sqlalchemy import func

from aegisScout.core.config import settings
from aegisScout.core import database as db_module
from aegisScout.core.database import init_db

class _EngineProxy:
    def __getattr__(self, name):
        return getattr(db_module.engine, name)

engine = _EngineProxy()
from aegisScout.core.models import (
    Lead,
    ResearchNote,
    Message,
    ActivityLog,
    Campaign,
    UserSession,
    SearchPreset,
    DiscoveryDraft,
    SmtpAccount,
)
from aegisScout.cli import commands
from aegisScout.outreach.assisted_mode import send_assisted_message
from aegisScout.gui_assets import HTML_CONTENT
from aegisScout.utils.logger import get_logger

logger = get_logger("gui.api")


def _utcnow() -> datetime:
    """Return current UTC datetime as timezone-naive (SQLite-safe)."""
    return datetime.now(timezone.utc).replace(tzinfo=None)


# ---------------------------------------------------------------------------
# JS bridge key classification
# ---------------------------------------------------------------------------
# Keys the JS layer may LEGITIMATELY set via `set_config_value`. Sensitive
# keys (api_key/password/token) are routed to .env via the Settings singleton;
# non-sensitive keys are routed to config.toml. Any other key is rejected.
_TOML_KEYS = {
    # path (dotted) -> TOML file section/key
    "outreach.language": ("outreach", "language"),
    "outreach.tone": ("outreach", "tone"),
    "outreach.company_name": ("outreach", "company_name"),
    "outreach.mod_b_acknowledged": ("outreach", "mod_b_acknowledged"),
    "rate_limits.instagram_actions_per_hour": ("rate_limits", "instagram_actions_per_hour"),
    "rate_limits.instagram_actions_per_day": ("rate_limits", "instagram_actions_per_day"),
    "rate_limits.scrape_delay_min_seconds": ("rate_limits", "scrape_delay_min_seconds"),
    "rate_limits.scrape_delay_max_seconds": ("rate_limits", "scrape_delay_max_seconds"),
    "discovery.default_search_radius_km": ("discovery", "default_search_radius_km"),
    "discovery.target_sectors": ("discovery", "target_sectors"),
    "app.name": ("app", "name"),
    "notifications.telegram_enabled": ("notifications", "telegram_enabled"),
    "notifications.email_enabled": ("notifications", "email_enabled"),
    "app.show_console": ("app", "show_console"),
}

# Sensitive keys — must be persisted to .env, never to config.toml.
_ENV_KEYS = {
    "llm_primary_provider",
    "llm_fallback_provider",
    "deepseek_api_key",
    "anthropic_api_key",
    "openai_api_key",
    "openrouter_api_key",
    "gemini_api_key",
    "groq_api_key",
    "mistral_api_key",
    "ollama_base_url",
    "llm_timeout",
    "discovery_primary_provider",
    "google_places_api_key",
    "google_custom_search_api_key",
    "google_custom_search_cx",
    "instagram_username",
    "instagram_session_encryption_key",
    "telegram_bot_token",
    "telegram_chat_id",
    "notify_email_smtp_host",
    "notify_email_smtp_port",
    "notify_email_username",
    "notify_email_password",
    "notify_email_imap_host",
    "notify_email_imap_port",
    "notify_email_imap_username",
    "notify_email_imap_password",
    "proxy_pool",
    "scraper_api_key",
    "zenrows_api_key",
    "crawlbase_api_key",
    "apify_api_key",
    "linkedin_session_cookie",
    "ollama_model",
    "outreach_mode",
    "max_daily_outreach",
}

# Whitelist of safe commands that trigger_command() will dispatch.
_SAFE_COMMANDS = {
    "list",
    "discover",
    "research",
    "export",
    "is_configured",
    "get_stats",
    "get_leads",
    "get_lead_details",
    "get_campaigns",
    "get_campaign_details",
    "get_activity_logs",
    "get_sessions",
}


class GuiApi:
    """Python API exposed to the PyWebView JS frontend."""

    def __init__(self):
        init_db()
        self._active_session_id = 1
        self._window = None
        # In-memory only — for the login_instagram() flow. Never persisted
        # beyond the GUI process. Reset to None after a successful or failed
        # login to minimise accidental exposure.
        self._ig_session_password: str | None = None

    # -------------------------------------------------------------------
    # Session helpers (unchanged)
    # -------------------------------------------------------------------
    @property
    def active_session_id(self):
        return self._active_session_id

    def _resolve_session(self, session_id: int | None) -> int:
        return session_id if session_id is not None else self._active_session_id

    def _normalize_query(self, value) -> str:
        if value is None:
            return ""
        if isinstance(value, (list, tuple, set)):
            parts = [str(item).strip() for item in value if str(item).strip()]
            return ", ".join(parts)
        return str(value).strip()

    def _normalize_list_query(self, value) -> list[str]:
        normalized = self._normalize_query(value)
        if not normalized:
            return []
        return [part.strip() for part in re.split(r"[\n,;]", normalized) if part.strip()]

    def _resolve_radius(self, radius_km) -> int:
        """Resolve a radius_km parameter to a non-negative int.

        0 (or any non-numeric / empty input) means "unlimited" — the
        discovery command should skip the location-radius filter.
        WebSearch is unaffected; OSM accepts 0/None as no-radius; Google
        Places treats 0 as "skip radius constraint" (use max allowed).
        """
        if radius_km is None or radius_km == "":
            return 0
        try:
            parsed = int(radius_km)
        except (ValueError, TypeError):
            return 0
        return max(0, parsed)

    # -------------------------------------------------------------------
    # Leads (unchanged)
    # -------------------------------------------------------------------
    def get_leads(self, status=None, search_log_id=None, has_website="", has_instagram="", has_phone=""):
        try:
            with Session(engine) as session:
                stmt = select(Lead).where(Lead.session_id == self._active_session_id)
                if status and status != "all":
                    stmt = stmt.where(Lead.status == status)

                if has_website == "yes":
                    stmt = stmt.where(Lead.website_url.is_not(None))
                elif has_website == "no":
                    stmt = stmt.where(Lead.website_url.is_(None))

                if has_instagram == "yes":
                    stmt = stmt.where(Lead.instagram_handle.is_not(None))
                elif has_instagram == "no":
                    stmt = stmt.where(Lead.instagram_handle.is_(None))

                if has_phone == "yes":
                    stmt = stmt.where(Lead.phone.is_not(None), Lead.phone != "")
                elif has_phone == "no":
                    stmt = stmt.where((Lead.phone.is_(None)) | (Lead.phone == ""))

                if search_log_id:
                    log = session.get(ActivityLog, int(search_log_id))
                    if log:
                        t_start = log.timestamp - timedelta(seconds=5)
                        t_end = log.timestamp + timedelta(seconds=5)
                        stmt = stmt.where(
                            Lead.discovered_at >= t_start,
                            Lead.discovered_at <= t_end
                        )

                stmt = stmt.order_by(Lead.id.desc())
                leads = session.exec(stmt).all()

                res = []
                for lead in leads:
                    ld = lead.model_dump()
                    if ld.get("discovered_at"):
                        ld["discovered_at"] = ld["discovered_at"].strftime("%Y-%m-%d %H:%M:%S")
                    if ld.get("updated_at"):
                        ld["updated_at"] = ld["updated_at"].strftime("%Y-%m-%d %H:%M:%S")
                    res.append(ld)
                return res
        except Exception as e:
            return {"error": str(e)}

    def get_search_history(self):
        try:
            with Session(engine) as session:
                stmt = (
                    select(ActivityLog)
                    .where(ActivityLog.session_id == self._active_session_id)
                    .where(ActivityLog.action == "discover")
                )
                if hasattr(ActivityLog, "deleted"):
                    stmt = stmt.where(ActivityLog.deleted != 1)
                if hasattr(ActivityLog, "status"):
                    stmt = stmt.where(ActivityLog.status != "deleted")
                stmt = stmt.order_by(ActivityLog.timestamp.desc())
                logs = session.exec(stmt).all()
                
                pattern = r"Discovered (\d+) new leads(?:\s*\(skipped (\d+) duplicates\)\s*out of (\d+) candidates)? for sector='(.*?)', location='(.*?)' via '(.*?)'"
                
                results = []
                for log in logs:
                    details = log.details or ""
                    m = re.match(pattern, details)
                    if m:
                        added = int(m.group(1))
                        skipped = int(m.group(2)) if m.group(2) is not None else 0
                        total = int(m.group(3)) if m.group(3) is not None else added
                        sector = m.group(4)
                        location = m.group(5)
                        provider = m.group(6)
                    else:
                        added = 0
                        skipped = 0
                        total = 0
                        sector = "Bilinmeyen"
                        location = "Bilinmeyen"
                        provider = "osm"
                        
                        if "sector=" in details:
                            sec_m = re.search(r"sector='(.*?)'", details)
                            if sec_m: sector = sec_m.group(1)
                            loc_m = re.search(r"location='(.*?)'", details)
                            if loc_m: location = loc_m.group(1)
                            prov_m = re.search(r"via '(.*?)'", details)
                            if prov_m: provider = prov_m.group(1)
                            cnt_m = re.search(r"Discovered (\d+)", details)
                            if cnt_m: added = int(cnt_m.group(1))
                    
                    results.append({
                        "id": log.id,
                        "timestamp": log.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                        "sector": sector,
                        "location": location,
                        "provider": provider,
                        "added_count": added,
                        "duplicate_count": skipped,
                        "total_candidates": total,
                        "details": details
                    })
                return results
        except Exception as e:
            return {"error": str(e)}

    def delete_leads_by_search(self, search_log_id):
        try:
            with Session(engine) as session:
                log = session.get(ActivityLog, int(search_log_id))
                if not log:
                    return {"error": "Arama kaydı bulunamadı."}
                
                t_start = log.timestamp - timedelta(seconds=5)
                t_end = log.timestamp + timedelta(seconds=5)
                
                leads_stmt = select(Lead).where(
                    Lead.session_id == self._active_session_id,
                    Lead.discovered_at >= t_start,
                    Lead.discovered_at <= t_end
                )
                leads = session.exec(leads_stmt).all()
                deleted_count = len(leads)
                
                for lead in leads:
                    session.delete(lead)
                
                session.delete(log)
                
                log_entry = ActivityLog(
                    action="leads_clear_search",
                    details=f"Aramaya ait adaylar temizlendi (Arama ID: {search_log_id}, Silinen: {deleted_count}).",
                    session_id=self._active_session_id,
                )
                session.add(log_entry)
                session.commit()
                return {"success": True, "deleted_count": deleted_count}
        except Exception as e:
            return {"error": str(e)}

    def clear_leads_advanced(self, filters):
        try:
            with Session(engine) as session:
                stmt = select(Lead).where(Lead.session_id == self._active_session_id)
                
                status = filters.get("status")
                if status and status != "all":
                    stmt = stmt.where(Lead.status == status)
                    
                sector = filters.get("sector")
                if sector and sector.strip():
                    stmt = stmt.where(Lead.sector == sector.strip())
                    
                source = filters.get("source")
                if source and source != "all":
                    stmt = stmt.where(Lead.source == source)
                    
                min_rating = filters.get("min_rating")
                if min_rating is not None and str(min_rating).strip() != "":
                    try:
                        stmt = stmt.where(Lead.rating >= float(min_rating))
                    except ValueError:
                        pass
                max_rating = filters.get("max_rating")
                if max_rating is not None and str(max_rating).strip() != "":
                    try:
                        stmt = stmt.where(Lead.rating <= float(max_rating))
                    except ValueError:
                        pass
                        
                has_website = filters.get("has_website")
                if has_website == "yes":
                    stmt = stmt.where(Lead.has_website == True)
                elif has_website == "no":
                    stmt = stmt.where(Lead.has_website == False)
                    
                start_date = filters.get("start_date")
                if start_date and start_date.strip():
                    try:
                        dt_start = datetime.strptime(start_date, "%Y-%m-%d")
                        stmt = stmt.where(Lead.discovered_at >= dt_start)
                    except ValueError:
                        pass
                end_date = filters.get("end_date")
                if end_date and end_date.strip():
                    try:
                        dt_end = datetime.strptime(end_date + " 23:59:59", "%Y-%m-%d %H:%M:%S")
                        stmt = stmt.where(Lead.discovered_at <= dt_end)
                    except ValueError:
                        pass
                
                search_log_id = filters.get("search_log_id")
                if search_log_id and str(search_log_id).strip() != "":
                    log = session.get(ActivityLog, int(search_log_id))
                    if log:
                        t_start = log.timestamp - timedelta(seconds=5)
                        t_end = log.timestamp + timedelta(seconds=5)
                        stmt = stmt.where(
                            Lead.discovered_at >= t_start,
                            Lead.discovered_at <= t_end
                        )
                
                keyword = filters.get("keyword")
                if keyword and keyword.strip():
                    kw_lower = keyword.strip().lower()
                    stmt = stmt.where(
                        func.lower(Lead.business_name).contains(kw_lower) |
                        func.lower(Lead.sector).contains(kw_lower) |
                        func.lower(Lead.address).contains(kw_lower) |
                        func.lower(Lead.website_url).contains(kw_lower)
                    )
                
                leads = session.exec(stmt).all()
                count = len(leads)
                
                if filters.get("dry_run", False):
                    return {"success": True, "count": count}
                
                for lead in leads:
                    session.delete(lead)
                
                log_entry = ActivityLog(
                    action="leads_clear_advanced",
                    details=f"Gelişmiş aday temizleme yapıldı. Silinen aday: {count}.",
                    session_id=self._active_session_id
                )
                session.add(log_entry)
                session.commit()
                return {"success": True, "deleted_count": count}
        except Exception as e:
            return {"error": str(e)}

    def export_leads_dialog(self, filters):
        if not self._window:
            return {"error": "GUI penceresi hazır değil."}
            
        try:
            with Session(engine) as session:
                stmt = select(Lead).where(Lead.session_id == self._active_session_id)
                
                status = filters.get("status")
                if status and status != "all":
                    stmt = stmt.where(Lead.status == status)
                    
                sector = filters.get("sector")
                if sector and sector.strip():
                    stmt = stmt.where(Lead.sector == sector.strip())
                    
                search_log_id = filters.get("search_log_id")
                if search_log_id and str(search_log_id).strip() != "":
                    log = session.get(ActivityLog, int(search_log_id))
                    if log:
                        t_start = log.timestamp - timedelta(seconds=5)
                        t_end = log.timestamp + timedelta(seconds=5)
                        stmt = stmt.where(
                            Lead.discovered_at >= t_start,
                            Lead.discovered_at <= t_end
                        )
                        
                keyword = filters.get("keyword")
                if keyword and keyword.strip():
                    kw_lower = keyword.strip().lower()
                    stmt = stmt.where(
                        func.lower(Lead.business_name).contains(kw_lower) |
                        func.lower(Lead.sector).contains(kw_lower) |
                        func.lower(Lead.address).contains(kw_lower) |
                        func.lower(Lead.website_url).contains(kw_lower)
                    )

                leads = session.exec(stmt).all()
                if not leads:
                    return {"error": "Dışa aktarılacak aday bulunamadı."}
                    
                now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                det = "tumu"
                if status and status != "all":
                    det = f"durum_{status}"
                elif sector and sector.strip():
                    det = f"sektor_{sector.strip()}"
                elif search_log_id and str(search_log_id).strip() != "":
                    det = f"arama_{search_log_id}"
                    
                det = re.sub(r"[^\w\-]", "_", det)
                default_filename = f"aegisScout_adaylar_{det}_{now_str}"
                
                file_types = (
                    "Excel Dosyası (*.xlsx)",
                    "CSV Dosyası (*.csv)",
                    "PDF Raporu (*.pdf)",
                    "JSON Verisi (*.json)",
                    "HTML Sayfası (*.html)",
                    "Tüm Dosyalar (*.*)"
                )
                
                result = self._window.create_file_dialog(
                    webview.SAVE_DIALOG,
                    save_filename=default_filename,
                    file_types=file_types
                )
                
                if not result:
                    return {"success": False, "cancelled": True}
                    
                save_path = result[0]
                ext = Path(save_path).suffix.lower()
                
                # Default format mapping based on UI option selection
                ui_format = filters.get("format", ".xlsx")
                if not ext:
                    save_path += ui_format
                    ext = ui_format
                    
                if ext == ".csv":
                    import csv
                    with open(save_path, "w", newline="", encoding="utf-8-sig") as f:
                        writer = csv.writer(f)
                        writer.writerow([
                            "ID", "İşletme Adı", "Sektör", "Telefon", "Adres", "Web Sitesi",
                            "Instagram", "YouTube", "LinkedIn", "TikTok", "Facebook",
                            "Telegram", "X/Twitter", "Puan", "Yorum Sayısı", "Web Skor",
                            "Durum", "Keşif Tarihi"
                        ])
                        for lead in leads:
                            writer.writerow([
                                lead.id, lead.business_name, lead.sector or "", lead.phone or "",
                                lead.address or "", lead.website_url or "", lead.instagram_handle or "",
                                lead.youtube_url or "", lead.linkedin_url or "",
                                lead.tiktok_url or "", lead.facebook_url or "",
                                lead.telegram_url or "", lead.twitter_url or "",
                                lead.rating or "", lead.review_count or "", lead.website_quality_score or "",
                                lead.status, lead.discovered_at.strftime("%Y-%m-%d %H:%M:%S") if lead.discovered_at else ""
                            ])
                            
                elif ext == ".xlsx":
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Adaylar"
                    
                    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    align_center = Alignment(horizontal="center", vertical="center")
                    align_left = Alignment(horizontal="left", vertical="center")
                    
                    headers = [
                        "ID", "İşletme Adı", "Sektör", "Telefon", "Adres", "Web Sitesi",
                        "Instagram", "YouTube", "LinkedIn", "TikTok", "Facebook",
                        "Telegram", "X/Twitter", "Puan", "Yorum Sayısı", "Web Skor",
                        "Durum", "Keşif Tarihi"
                    ]
                    ws.append(headers)
                    
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=1, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = align_center if col_num in [1, 14, 15, 16, 17, 18] else align_left
                        
                    for lead in leads:
                        ws.append([
                            lead.id, lead.business_name, lead.sector or "", lead.phone or "",
                            lead.address or "", lead.website_url or "", lead.instagram_handle or "",
                            lead.youtube_url or "", lead.linkedin_url or "",
                            lead.tiktok_url or "", lead.facebook_url or "",
                            lead.telegram_url or "", lead.twitter_url or "",
                            lead.rating or "", lead.review_count or "", lead.website_quality_score or "",
                            lead.status, lead.discovered_at.strftime("%Y-%m-%d %H:%M:%S") if lead.discovered_at else ""
                        ])
                        
                    thin_border = Border(
                        left=Side(style='thin', color='D9D9D9'),
                        right=Side(style='thin', color='D9D9D9'),
                        top=Side(style='thin', color='D9D9D9'),
                        bottom=Side(style='thin', color='D9D9D9')
                    )
                    
                    for row in range(2, len(leads) + 2):
                        for col in range(1, len(headers) + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.font = Font(name="Segoe UI", size=10)
                            cell.border = thin_border
                            cell.alignment = align_center if col in [1, 8, 9, 10, 11, 12] else align_left
                            
                    for col in ws.columns:
                        max_len = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            val_str = str(cell.value or '')
                            if len(val_str) > max_len:
                                max_len = len(val_str)
                        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
                        
                    wb.save(save_path)
                    
                elif ext == ".json":
                    data = []
                    for lead in leads:
                        ld = lead.model_dump()
                        if ld.get("discovered_at"):
                            ld["discovered_at"] = ld["discovered_at"].isoformat()
                        if ld.get("updated_at"):
                            ld["updated_at"] = ld["updated_at"].isoformat()
                        data.append(ld)
                    with open(save_path, "w", encoding="utf-8") as f:
                        json.dump(data, f, ensure_ascii=False, indent=2)
                        
                elif ext == ".html":
                    html_rows = []
                    for lead in leads:
                        html_rows.append(f"""
                        <tr>
                            <td>{lead.id}</td>
                            <td class="bold">{lead.business_name}</td>
                            <td>{lead.sector or '-'}</td>
                            <td>{lead.phone or '-'}</td>
                            <td>{lead.address or '-'}</td>
                            <td><a href="{lead.website_url}" target="_blank">{lead.website_url or '-'}</a></td>
                            <td>{lead.instagram_handle or '-'}</td>
                            <td><a href="{lead.youtube_url}" target="_blank">{lead.youtube_url or '-'}</a></td>
                            <td><a href="{lead.linkedin_url}" target="_blank">{lead.linkedin_url or '-'}</a></td>
                            <td><a href="{lead.tiktok_url}" target="_blank">{lead.tiktok_url or '-'}</a></td>
                            <td><a href="{lead.facebook_url}" target="_blank">{lead.facebook_url or '-'}</a></td>
                            <td><a href="{lead.telegram_url}" target="_blank">{lead.telegram_url or '-'}</a></td>
                            <td><a href="{lead.twitter_url}" target="_blank">{lead.twitter_url or '-'}</a></td>
                            <td>{lead.rating or '-'}</td>
                            <td>{lead.review_count or '-'}</td>
                            <td>{lead.website_quality_score or '-'}</td>
                            <td><span class="badge status-{lead.status}">{lead.status}</span></td>
                            <td>{lead.discovered_at.strftime("%Y-%m-%d %H:%M:%S") if lead.discovered_at else '-'}</td>
                        </tr>
                        """)
                    
                    html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>aegisScout Aday Raporu</title>
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f8f9fa; margin: 0; padding: 24px; color: #333; }}
        .header {{ margin-bottom: 24px; border-bottom: 2px solid #1f4e78; padding-bottom: 12px; }}
        h1 {{ margin: 0; color: #1f4e78; font-size: 24px; }}
        .meta {{ color: #666; font-size: 14px; margin-top: 4px; }}
        table {{ width: 100%; border-collapse: collapse; background: #fff; margin-top: 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
        th, td {{ padding: 10px 12px; text-align: left; border-bottom: 1px solid #dee2e6; font-size: 13px; }}
        th {{ background-color: #1f4e78; color: white; font-weight: 600; }}
        tr:hover {{ background-color: #f1f3f5; }}
        .bold {{ font-weight: bold; color: #111; }}
        .badge {{ padding: 4px 8px; border-radius: 4px; font-size: 11px; font-weight: bold; text-transform: uppercase; }}
        .status-new {{ background-color: #e3f2fd; color: #0d47a1; }}
        .status-researched {{ background-color: #e8f5e9; color: #1b5e20; }}
        .status-contacted {{ background-color: #fff3e0; color: #e65100; }}
        .status-replied {{ background-color: #f3e5f5; color: #4a148c; }}
        .status-rejected {{ background-color: #ffebee; color: #c62828; }}
        a {{ color: #1f4e78; text-decoration: none; }}
        a:hover {{ text-decoration: underline; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>aegisScout Aday Keşif Raporu</h1>
        <div class="meta">Oluşturulma Tarihi: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | Toplam Kayıt: {len(leads)}</div>
    </div>
    <table>
        <thead>
            <tr>
                <th>ID</th>
                <th>İşletme Adı</th>
                <th>Sektör</th>
                <th>Telefon</th>
                <th>Adres</th>
                <th>Web Sitesi</th>
                <th>Instagram</th>
                <th>YouTube</th>
                <th>LinkedIn</th>
                <th>TikTok</th>
                <th>Facebook</th>
                <th>Telegram</th>
                <th>X/Twitter</th>
                <th>Puan</th>
                <th>Yorum</th>
                <th>Web Skor</th>
                <th>Durum</th>
                <th>Keşif Tarihi</th>
            </tr>
        </thead>
        <tbody>
            {"".join(html_rows)}
        </tbody>
    </table>
</body>
</html>
"""
                    with open(save_path, "w", encoding="utf-8") as f:
                        f.write(html_content)
                        
                elif ext == ".pdf":
                    from reportlab.lib.pagesizes import letter, landscape
                    from reportlab.lib import colors
                    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
                    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                    
                    doc = SimpleDocTemplate(save_path, pagesize=landscape(letter), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
                    story = []
                    
                    styles = getSampleStyleSheet()
                    title_style = ParagraphStyle(
                        'TitleStyle',
                        parent=styles['Heading1'],
                        fontName='Helvetica-Bold',
                        fontSize=18,
                        textColor=colors.HexColor('#1F4E78'),
                        spaceAfter=4
                    )
                    meta_style = ParagraphStyle(
                        'MetaStyle',
                        parent=styles['Normal'],
                        fontName='Helvetica',
                        fontSize=10,
                        textColor=colors.HexColor('#666666'),
                        spaceAfter=15
                    )
                    cell_style = ParagraphStyle(
                        'CellStyle',
                        parent=styles['Normal'],
                        fontName='Helvetica',
                        fontSize=7,
                        leading=9
                    )
                    cell_bold_style = ParagraphStyle(
                        'CellBoldStyle',
                        parent=cell_style,
                        fontName='Helvetica-Bold'
                    )
                    
                    story.append(Paragraph("aegisScout — Business Discovery Report", title_style))
                    story.append(Paragraph(f"Created: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Leads: {len(leads)}", meta_style))
                    
                    data = [[
                        Paragraph("<b>Name</b>", cell_bold_style),
                        Paragraph("<b>Sector</b>", cell_bold_style),
                        Paragraph("<b>Phone</b>", cell_bold_style),
                        Paragraph("<b>Address</b>", cell_bold_style),
                        Paragraph("<b>Website</b>", cell_bold_style),
                        Paragraph("<b>Instagram</b>", cell_bold_style),
                        Paragraph("<b>Rating</b>", cell_bold_style),
                        Paragraph("<b>Reviews</b>", cell_bold_style),
                        Paragraph("<b>Web Score</b>", cell_bold_style),
                        Paragraph("<b>Status</b>", cell_bold_style),
                    ]]
                    
                    for lead in leads:
                        data.append([
                            Paragraph(lead.business_name, cell_bold_style),
                            Paragraph(lead.sector or "", cell_style),
                            Paragraph(lead.phone or "", cell_style),
                            Paragraph(lead.address or "", cell_style),
                            Paragraph(lead.website_url or "", cell_style),
                            Paragraph(lead.instagram_handle or "", cell_style),
                            Paragraph(str(lead.rating) if lead.rating is not None else "", cell_style),
                            Paragraph(str(lead.review_count) if lead.review_count is not None else "", cell_style),
                            Paragraph(str(lead.website_quality_score) if lead.website_quality_score is not None else "", cell_style),
                            Paragraph(lead.status, cell_bold_style),
                        ])
                    
                    col_widths = [120, 60, 65, 125, 120, 60, 30, 35, 45, 45]
                    t = Table(data, colWidths=col_widths, repeatRows=1)
                    t.setStyle(TableStyle([
                        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E78')),
                        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                        ('BOTTOMPADDING', (0,0), (-1,0), 6),
                        ('TOPPADDING', (0,0), (-1,0), 6),
                        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#F2F2F2'), colors.white]),
                        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')),
                        ('VALIGN', (0,0), (-1,-1), 'TOP'),
                    ]))
                    story.append(t)
                    doc.build(story)
                
                log = ActivityLog(
                    action="export",
                    details=f"Dışa aktarma tamamlandı ({len(leads)} kayıt, Dosya: {os.path.basename(save_path)}, Format: {ext}).",
                    session_id=self._active_session_id
                )
                session.add(log)
                session.commit()
                
                return {"success": True, "path": save_path, "count": len(leads)}
        except Exception as e:
            logger.error(f"Dışa aktarma hatası: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return {"error": str(e)}

    def get_lead_details(self, lead_id):
        try:
            with Session(engine) as session:
                lead = session.get(Lead, lead_id)
                if not lead:
                    return {"error": f"Lead {lead_id} not found."}

                notes = [note.model_dump() for note in lead.research_notes]
                messages = [msg.model_dump() for msg in lead.messages]

                for n in notes:
                    n["scraped_at"] = n["scraped_at"].strftime("%Y-%m-%d %H:%M:%S")
                for m in messages:
                    m["created_at"] = m["created_at"].strftime("%Y-%m-%d %H:%M:%S")
                    if m["sent_at"]:
                        m["sent_at"] = m["sent_at"].strftime("%Y-%m-%d %H:%M:%S")

                lead_data = lead.model_dump()
                lead_data["discovered_at"] = lead_data["discovered_at"].strftime("%Y-%m-%d %H:%M:%S")
                lead_data["updated_at"] = lead_data["updated_at"].strftime("%Y-%m-%d %H:%M:%S")

                return {
                    "lead": lead_data,
                    "notes": notes,
                    "messages": messages,
                }
        except Exception as e:
            return {"error": str(e)}

    def discover_leads(self, sector, location, radius, provider):
        try:
            parsed_radius = self._resolve_radius(radius)
            if parsed_radius == 0:
                logger.info(
                    "discover_leads: radius=unlimited (0/empty) — "
                    "discovery will run without a location-radius filter."
                )

            sector_query = self._normalize_query(sector)
            location_query = self._normalize_query(location)

            task_id = f"discovery_{int(time.time())}"
            task_name = f"Tarama: {sector} ({location})"

            async def task_coro(task_id: str):
                try:
                    def progress_cb(msg):
                        if self._window:
                            try:
                                escaped_msg = json.dumps(msg)
                                self._window.evaluate_js(
                                    f"updateDiscoveryProgress({escaped_msg})"
                                )
                            except Exception:
                                pass

                    res = await commands.discover_leads(
                        sector=sector_query,
                        location=location_query,
                        radius_km=parsed_radius,
                        provider_name=provider,
                        progress_callback=progress_cb,
                        session_id=self._active_session_id,
                        task_id=task_id
                    )
                    added_count, total_candidates, duplicate_count = res
                    if self._window:
                        try:
                            self._window.evaluate_js(
                                f"finishDiscovery(true, {added_count}, null, {total_candidates}, {duplicate_count})"
                            )
                        except Exception:
                            pass
                except Exception as ex:
                    logger.error(f"Discovery task failed: {ex}")
                    if self._window:
                        try:
                            err_json = json.dumps(str(ex))
                            self._window.evaluate_js(
                                f"finishDiscovery(false, 0, {err_json}, 0, 0)"
                            )
                        except Exception:
                            pass

            from aegisScout.core.task_queue import TaskQueueManager
            tqm = TaskQueueManager.get_instance()
            tqm.add_task(task_id, task_name, task_coro)
            return {"success": True, "queued": True, "task_id": task_id}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def list_search_presets(self):
        try:
            with Session(engine) as session:
                presets = session.exec(
                    select(SearchPreset)
                    .where(SearchPreset.session_id == self._active_session_id)
                    .order_by(SearchPreset.id.desc())
                ).all()
                return [preset.model_dump() for preset in presets]
        except Exception as e:
            return {"error": str(e)}

    def save_search_preset(self, name, sector, location, radius, provider, notes=None):
        try:
            if not name or not str(name).strip():
                return {"error": "Preset name is required."}
            with Session(engine) as session:
                preset = SearchPreset(
                    name=str(name).strip(),
                    sector_query=self._normalize_query(sector),
                    location_query=self._normalize_query(location),
                    radius_km=int(radius) if radius not in (None, "") else 10,
                    provider_name=str(provider or "all").strip(),
                    notes=self._normalize_query(notes) or None,
                    session_id=self._active_session_id,
                )
                session.add(preset)
                session.commit()
                session.refresh(preset)
                return {"success": True, "preset": preset.model_dump()}
        except Exception as e:
            return {"error": str(e)}

    def list_search_drafts(self):
        try:
            with Session(engine) as session:
                drafts = session.exec(
                    select(DiscoveryDraft)
                    .where(DiscoveryDraft.session_id == self._active_session_id)
                    .order_by(DiscoveryDraft.id.desc())
                ).all()
                return [draft.model_dump() for draft in drafts]
        except Exception as e:
            return {"error": str(e)}

    def save_search_draft(self, name, sector, location, radius, provider, country=None, city=None, region=None, keywords=None, notes=None):
        try:
            if not name or not str(name).strip():
                return {"error": "Draft name is required."}
            with Session(engine) as session:
                draft = DiscoveryDraft(
                    name=str(name).strip(),
                    sector_query=self._normalize_query(sector),
                    location_query=self._normalize_query(location),
                    radius_km=int(radius) if radius not in (None, "") else 10,
                    provider_name=str(provider or "all").strip(),
                    country_query=self._normalize_query(country) or None,
                    city_query=self._normalize_query(city) or None,
                    region_query=self._normalize_query(region) or None,
                    keyword_query=self._normalize_query(keywords) or None,
                    notes=self._normalize_query(notes) or None,
                    session_id=self._active_session_id,
                )
                session.add(draft)
                session.commit()
                session.refresh(draft)
                return {"success": True, "draft": draft.model_dump()}
        except Exception as e:
            return {"error": str(e)}

    def load_search_draft(self, draft_id):
        try:
            draft_id = int(draft_id)
            with Session(engine) as session:
                draft = session.get(DiscoveryDraft, draft_id)
                if not draft:
                    return {"error": "Draft not found."}
                return {"success": True, "draft": draft.model_dump()}
        except Exception as e:
            return {"error": str(e)}

    def research_lead(self, lead_id, force: bool = True):
        try:
            lead_id = int(lead_id)
            task_id = f"research_{lead_id}_{int(time.time())}"
            
            lead_name = f"ID: {lead_id}"
            with Session(db_module.engine) as session:
                lead = session.get(Lead, lead_id)
                if lead:
                    lead_name = lead.business_name
            task_name = f"Araştırma: {lead_name}"

            async def task_coro(task_id: str):
                try:
                    await commands.research_lead(lead_id, force=force, task_id=task_id)
                    if self._window:
                        try:
                            self._window.evaluate_js(
                                f"finishResearch({lead_id}, true, null)"
                            )
                        except Exception:
                            pass
                except Exception as ex:
                    logger.error(f"Research task failed: {ex}")
                    if self._window:
                        try:
                            err_json = json.dumps(str(ex))
                            self._window.evaluate_js(
                                f"finishResearch({lead_id}, false, {err_json})"
                            )
                        except Exception:
                            pass

            from aegisScout.core.task_queue import TaskQueueManager
            tqm = TaskQueueManager.get_instance()
            tqm.add_task(task_id, task_name, task_coro)
            return {"success": True, "queued": True, "task_id": task_id}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def save_message_draft(self, lead_id, content):
        try:
            with Session(engine) as session:
                lead = session.get(Lead, lead_id)
                if not lead:
                    return {"error": "Lead not found."}

                old_drafts = session.exec(
                    select(Message).where(
                        (Message.lead_id == lead.id) & (Message.status == "draft")
                    )
                ).all()
                for d in old_drafts:
                    session.delete(d)

                new_msg = Message(
                    lead_id=lead.id,
                    direction="outbound",
                    channel="instagram_manual",
                    content=content,
                    ai_generated=False,
                    status="draft",
                )
                session.add(new_msg)
                session.commit()
                return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    # -------------------------------------------------------------------
    # Send operations (unchanged / expanded with Email & WhatsApp)
    # -------------------------------------------------------------------
    def send_email_lead(self, lead_id, subject, body):
        try:
            with Session(engine) as session:
                lead = session.get(Lead, lead_id)
                if not lead:
                    return {"error": "Aday bulunamadı."}
                if not lead.email:
                    return {"error": "Adayın e-posta adresi bulunamadı."}
                
                from aegisScout.outreach.email_client import send_cold_email
                success, err, used_smtp_id = send_cold_email(lead.email, subject, body)
                if success:
                    # Log email message
                    new_msg = Message(
                        lead_id=lead.id,
                        direction="outbound",
                        channel="email",
                        content=body,
                        status="sent",
                        sent_at=_utcnow(),
                        smtp_account_id=used_smtp_id
                    )
                    lead.status = "contacted"
                    lead.updated_at = _utcnow()
                    session.add(new_msg)
                    session.add(lead)
                    session.commit()
                    return {"success": True}
                else:
                    return {"error": err}
        except Exception as e:
            return {"error": str(e)}

    # -------------------------------------------------------------------
    # Task Queue APIs
    # -------------------------------------------------------------------
    def get_background_tasks(self):
        try:
            from aegisScout.core.task_queue import TaskQueueManager
            return {"success": True, "tasks": TaskQueueManager.get_instance().get_all_statuses()}
        except Exception as e:
            return {"error": str(e)}

    def pause_background_task(self, task_id):
        try:
            from aegisScout.core.task_queue import TaskQueueManager
            success = TaskQueueManager.get_instance().pause_task(task_id)
            return {"success": success}
        except Exception as e:
            return {"error": str(e)}

    def resume_background_task(self, task_id):
        try:
            from aegisScout.core.task_queue import TaskQueueManager
            success = TaskQueueManager.get_instance().resume_task(task_id)
            return {"success": success}
        except Exception as e:
            return {"error": str(e)}

    def cancel_background_task(self, task_id):
        try:
            from aegisScout.core.task_queue import TaskQueueManager
            success = TaskQueueManager.get_instance().cancel_task(task_id)
            return {"success": success}
        except Exception as e:
            return {"error": str(e)}

    def run_deep_osint_scan(self, lead_id: int):
        """Execute async Deep OSINT scan for a target lead."""
        import asyncio
        try:
            from aegisScout.core.deep_osint import DeepOSINTScanner
            scanner = DeepOSINTScanner()
            result = asyncio.run(scanner.scan_lead(lead_id))
            return {"success": True, "osint": result}
        except Exception as e:
            return {"error": str(e)}

    # -------------------------------------------------------------------
    # SMTP Account APIs
    # -------------------------------------------------------------------
    def get_smtp_accounts(self):
        try:
            with Session(engine) as session:
                accounts = session.exec(select(SmtpAccount)).all()
                return {"success": True, "accounts": [acc.model_dump() for acc in accounts]}
        except Exception as e:
            return {"error": str(e)}

    def add_smtp_account(self, name, host, port, user, password, imap_host=None, imap_port=None, imap_user=None, imap_pass=None):
        try:
            from aegisScout.utils.encryption import encrypt_string
            
            enc_pass = encrypt_string(password)
            enc_imap_pass = encrypt_string(imap_pass) if imap_pass else enc_pass
            
            with Session(engine) as session:
                new_acc = SmtpAccount(
                    name=name,
                    smtp_host=host,
                    smtp_port=int(port) if port else 587,
                    smtp_user=user,
                    smtp_pass=enc_pass,
                    imap_host=imap_host,
                    imap_port=int(imap_port) if imap_port else 993,
                    imap_user=imap_user or user,
                    imap_pass=enc_imap_pass,
                    is_active=True
                )
                session.add(new_acc)
                session.commit()
                return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    def toggle_smtp_account(self, acc_id, is_active):
        try:
            with Session(engine) as session:
                acc = session.get(SmtpAccount, int(acc_id))
                if not acc:
                    return {"error": "SMTP account not found."}
                acc.is_active = bool(is_active)
                session.add(acc)
                session.commit()
                return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    def delete_smtp_account(self, acc_id):
        try:
            with Session(engine) as session:
                acc = session.get(SmtpAccount, int(acc_id))
                if not acc:
                    return {"error": "SMTP account not found."}
                session.delete(acc)
                session.commit()
                return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    # -------------------------------------------------------------------
    # Campaign Follow-up Settings APIs
    # -------------------------------------------------------------------
    def get_campaign_followup_settings(self, campaign_id):
        try:
            with Session(engine) as session:
                campaign = session.get(Campaign, int(campaign_id))
                if not campaign:
                    return {"error": "Campaign not found."}
                return {
                    "success": True,
                    "settings": {
                        "followup_delay_1_days": campaign.followup_delay_1_days,
                        "followup_subject_1": campaign.followup_subject_1,
                        "followup_body_1": campaign.followup_body_1,
                        "followup_delay_2_days": campaign.followup_delay_2_days,
                        "followup_subject_2": campaign.followup_subject_2,
                        "followup_body_2": campaign.followup_body_2,
                    }
                }
        except Exception as e:
            return {"error": str(e)}

    def save_campaign_followup_settings(self, campaign_id, delay1, subject1, body1, delay2, subject2, body2):
        try:
            with Session(engine) as session:
                campaign = session.get(Campaign, int(campaign_id))
                if not campaign:
                    return {"error": "Campaign not found."}
                campaign.followup_delay_1_days = int(delay1) if delay1 else None
                campaign.followup_subject_1 = subject1
                campaign.followup_body_1 = body1
                campaign.followup_delay_2_days = int(delay2) if delay2 else None
                campaign.followup_subject_2 = subject2
                campaign.followup_body_2 = body2
                session.add(campaign)
                session.commit()
                return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    # -------------------------------------------------------------------
    # RAG / Knowledge Base APIs
    # -------------------------------------------------------------------
    def get_kb_files(self):
        try:
            from aegisScout.ai.local_rag import get_all_kb_files
            files = get_all_kb_files()
            return {"success": True, "files": [f.name for f in files]}
        except Exception as e:
            return {"error": str(e)}

    def upload_kb_file(self, file_name, file_content):
        try:
            import base64
            from aegisScout.ai.local_rag import KB_DIR, index_knowledge_base
            KB_DIR.mkdir(parents=True, exist_ok=True)
            
            try:
                decoded = base64.b64decode(file_content)
            except Exception:
                decoded = file_content.encode("utf-8")
                
            dest = KB_DIR / file_name
            dest.write_bytes(decoded)
            
            index_knowledge_base()
            return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    def delete_kb_file(self, file_name):
        try:
            from aegisScout.ai.local_rag import KB_DIR, index_knowledge_base
            dest = KB_DIR / file_name
            if dest.exists():
                dest.unlink()
                index_knowledge_base()
            return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    def reindex_kb(self):
        try:
            from aegisScout.ai.local_rag import index_knowledge_base
            res = index_knowledge_base()
            return {"success": True, "details": res}
        except Exception as e:
            return {"error": str(e)}

    def check_imap_replies_bridge(self):
        try:
            from aegisScout.outreach.email_client import check_imap_replies
            replied_emails = check_imap_replies()
            if not replied_emails:
                return {"replied_count": 0}
            
            updated_count = 0
            with Session(engine) as session:
                for email_addr in replied_emails:
                    stmt = select(Lead).where(func.lower(Lead.email) == email_addr.lower())
                    leads = session.exec(stmt).all()
                    for lead in leads:
                        if lead.status != "replied":
                            lead.status = "replied"
                            lead.updated_at = _utcnow()
                            # Insert inbound message log
                            new_msg = Message(
                                lead_id=lead.id,
                                direction="inbound",
                                channel="email",
                                content="E-posta üzerinden yanıt alındı.",
                                status="sent",
                                sent_at=_utcnow()
                            )
                            session.add(lead)
                            session.add(new_msg)
                            updated_count += 1
                session.commit()
            return {"replied_count": updated_count}
        except Exception as e:
            return {"error": str(e)}

    def launch_whatsapp(self, lead_id, message_text):
        try:
            with Session(engine) as session:
                lead = session.get(Lead, lead_id)
                if not lead:
                    return {"error": "Aday bulunamadı."}
                if not lead.phone:
                    return {"error": "Adayın telefon numarası bulunamadı."}
                
                # Clean phone number (keep digits only)
                clean_phone = "".join(filter(str.isdigit, lead.phone))
                # Add Turkish country prefix if it looks like local number (starts with 5 and is 10 chars)
                if len(clean_phone) == 10 and clean_phone.startswith("5"):
                    clean_phone = "90" + clean_phone
                    
                import urllib.parse
                import webbrowser
                import pyperclip
                
                # Copy message to clipboard for ease of pasting
                try:
                    pyperclip.copy(message_text)
                except Exception as cp_err:
                    logger.warning(f"Could not copy message to clipboard: {cp_err}")
                
                # WhatsApp Web Link Mod A
                wa_url = f"https://wa.me/{clean_phone}?text={urllib.parse.quote(message_text)}"
                webbrowser.open(wa_url)
                
                # Log message
                new_msg = Message(
                    lead_id=lead.id,
                    direction="outbound",
                    channel="whatsapp",
                    content=message_text,
                    status="sent",
                    sent_at=_utcnow()
                )
                lead.status = "contacted"
                lead.updated_at = _utcnow()
                session.add(new_msg)
                session.add(lead)
                session.commit()
                return {"success": True, "url": wa_url}
        except Exception as e:
            return {"error": str(e)}

    # -------------------------------------------------------------------
    # CRM Logs
    # -------------------------------------------------------------------
    def add_crm_log(self, lead_id, note_text):
        try:
            lead_id = int(lead_id)
            if not note_text or not note_text.strip():
                return {"error": "Not boş olamaz."}
            with Session(engine) as session:
                lead = session.get(Lead, lead_id)
                if not lead:
                    return {"error": "Aday bulunamadı."}
                from aegisScout.core.models import CrmLog
                new_log = CrmLog(
                    lead_id=lead_id,
                    note=note_text.strip()
                )
                session.add(new_log)
                session.commit()
                return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    def get_crm_logs(self, lead_id):
        try:
            lead_id = int(lead_id)
            with Session(engine) as session:
                from aegisScout.core.models import CrmLog
                stmt = select(CrmLog).where(CrmLog.lead_id == lead_id).order_by(text("created_at DESC"))
                logs = session.exec(stmt).all()
                serialized = []
                for log in logs:
                    serialized.append({
                        "id": log.id,
                        "note": log.note,
                        "created_at": log.created_at.isoformat() if log.created_at else None
                    })
                return {"logs": serialized}
        except Exception as e:
            return {"error": str(e)}

    # -------------------------------------------------------------------
    # Ollama Model Management
    # -------------------------------------------------------------------
    def get_ollama_status(self):
        try:
            import httpx
            base_url = settings.ollama_base_url.rstrip("/")
            resp = httpx.get(f"{base_url}/api/tags", timeout=3.0)
            if resp.status_code == 200:
                data = resp.json()
                models = [m["name"] for m in data.get("models", [])]
                return {"status": "running", "models": models}
            return {"status": "error", "message": f"Ollama HTTP {resp.status_code} döndürdü."}
        except Exception as e:
            return {"status": "offline", "message": f"Ollama sunucusuna bağlanılamadı ({settings.ollama_base_url}): {e}"}

    def pull_ollama_model(self, model_name):
        try:
            import threading
            import httpx
            def do_pull():
                try:
                    base_url = settings.ollama_base_url.rstrip("/")
                    resp = httpx.post(f"{base_url}/api/pull", json={"name": model_name}, timeout=300.0)
                    logger.info(f"Ollama pulled {model_name} status: {resp.status_code}")
                except Exception as ex:
                    logger.error(f"Failed to pull Ollama model {model_name}: {ex}")

            t = threading.Thread(target=do_pull, daemon=True)
            t.start()
            return {"success": True, "message": f"{model_name} arka planda indirilmeye başlandı."}
        except Exception as e:
            return {"error": str(e)}

    def send_assisted(self, lead_id):
        try:
            with Session(engine) as session:
                lead = session.get(Lead, lead_id)
                if not lead:
                    return {"error": "Lead not found."}

                msg_stmt = select(Message).where(
                    (Message.lead_id == lead.id) & (Message.status == "draft")
                )
                message = session.exec(msg_stmt).first()
                if not message:
                    return {"error": "Draft message not found. Please research or write draft first."}

                success = send_assisted_message(lead, message.content)
                if success:
                    message.status = "sent"
                    message.channel = "instagram_manual"
                    message.sent_at = _utcnow()
                    lead.status = "contacted"
                    lead.updated_at = _utcnow()
                    session.add(message)
                    session.add(lead)
                    session.commit()
                    return {"success": True}
                else:
                    return {"error": "Failed to copy clipboard or open profile page."}
        except Exception as e:
            return {"error": str(e)}

    def send_automated(self, lead_id):
        try:
            lead_id = int(lead_id)
            flag_file = Path("data/sessions/automation_authorized.flag")
            if not flag_file.exists():
                return {
                    "error": "Mod B (Tam Otomasyon) etkin değil. "
                             "Lütfen ayarlardan veya CLI'dan etkinleştirin."
                }

            with Session(engine) as session:
                lead = session.get(Lead, lead_id)
                if not lead:
                    return {"error": "Aday bulunamadı."}

                if not lead.instagram_handle:
                    return {"error": "Instagram adresi eksik."}

                msg_stmt = select(Message).where(
                    (Message.lead_id == lead.id) & (Message.status == "draft")
                )
                message = session.exec(msg_stmt).first()
                if not message:
                    return {"error": "Taslak mesaj bulunamadı."}

            def run_send():
                try:
                    with Session(engine) as session:
                        lead = session.get(Lead, lead_id)
                        msg_stmt = select(Message).where(
                            (Message.lead_id == lead.id) & (Message.status == "draft")
                        )
                        message = session.exec(msg_stmt).first()

                        from aegisScout.outreach.instagram_client import InstagramClient
                        ig = InstagramClient()
                        success = ig.send_direct_message(lead.instagram_handle, message.content)
                        if success:
                            message.status = "sent"
                            message.channel = "instagram_auto"
                            message.sent_at = _utcnow()
                            lead.status = "contacted"
                            lead.updated_at = _utcnow()
                            session.add(message)
                            session.add(lead)
                            session.commit()
                            if self._window:
                                try:
                                    self._window.evaluate_js(
                                        f"finishSendAutomated({lead_id}, true, null)"
                                    )
                                except Exception:
                                    pass
                        else:
                            raise Exception(
                                "Otomatik Instagram mesajı gönderilemedi. "
                                "Oturumunuzu veya limitleri kontrol edin."
                            )
                except Exception as ex:
                    logger.error(f"Error sending automated DM to lead {lead_id}: {ex}")
                    if self._window:
                        try:
                            err_json = json.dumps(str(ex))
                            self._window.evaluate_js(
                                f"finishSendAutomated({lead_id}, false, {err_json})"
                            )
                        except Exception:
                            pass

            t = threading.Thread(target=run_send, daemon=True)
            t.start()
            return {"queued": True}
        except Exception as e:
            return {"error": str(e)}

    # -------------------------------------------------------------------
    # Settings — SANITIZED. This is the security boundary.
    # -------------------------------------------------------------------
    
    # -------------------------------------------------------------------
    # Multichannel Outreach, Email Verification & Vision Audit Bridges
    # -------------------------------------------------------------------
    def send_whatsapp_assisted(self, lead_id):
        try:
            lead_id = int(lead_id)
            with Session(engine) as session:
                lead = session.get(Lead, lead_id)
                if not lead:
                    return {"error": "Aday bulunamadı."}
                if not lead.phone:
                    return {"error": "Telefon numarası bulunamadı."}

                msg_stmt = select(Message).where(
                    (Message.lead_id == lead.id) & (Message.status == "draft")
                )
                message = session.exec(msg_stmt).first()
                draft = message.content if message else f"Merhaba {lead.business_name}, web sitenizi inceledik..."

                from aegisScout.outreach.assisted_mode import send_whatsapp_assisted
                res = send_whatsapp_assisted(lead.phone, draft)
                if res.get("success"):
                    if message:
                        message.status = "sent"
                        message.channel = "whatsapp_manual"
                        message.sent_at = _utcnow()
                        session.add(message)
                    lead.status = "contacted"
                    lead.updated_at = _utcnow()
                    session.add(lead)
                    session.commit()
                return res
        except Exception as e:
            return {"error": str(e)}

    def send_linkedin_assisted(self, lead_id):
        try:
            lead_id = int(lead_id)
            with Session(engine) as session:
                lead = session.get(Lead, lead_id)
                if not lead:
                    return {"error": "Aday bulunamadı."}

                msg_stmt = select(Message).where(
                    (Message.lead_id == lead.id) & (Message.status == "draft")
                )
                message = session.exec(msg_stmt).first()
                draft = message.content if message else f"Merhaba {lead.business_name}, profilinizi inceledik..."

                from aegisScout.outreach.assisted_mode import send_linkedin_assisted
                res = send_linkedin_assisted(lead.business_name, draft, lead.website_url)
                if res.get("success"):
                    if message:
                        message.status = "sent"
                        message.channel = "linkedin_manual"
                        message.sent_at = _utcnow()
                        session.add(message)
                    lead.status = "contacted"
                    lead.updated_at = _utcnow()
                    session.add(lead)
                    session.commit()
                return res
        except Exception as e:
            return {"error": str(e)}

    def verify_email_local(self, email: str):
        try:
            from aegisScout.outreach.email_verifier import EmailVerifier
            verifier = EmailVerifier(check_smtp=True, timeout=5.0)
            return verifier.verify(email)
        except Exception as e:
            return {"error": str(e)}

    def run_vision_audit(self, url: str):
        try:
            import asyncio
            from aegisScout.ai.vision_audit import VisionAuditManager
            mgr = VisionAuditManager()
            loop = asyncio.new_event_loop()
            res = loop.run_until_complete(mgr.audit_url(url))
            loop.close()
            return res
        except Exception as e:
            return {"error": str(e)}

    def get_proxy_status(self):
        try:
            from aegisScout.utils.proxy_pool import ProxyPoolManager
            pool = ProxyPoolManager.get_instance()
            return {
                "total_configured": len(pool.proxies),
                "active_healthy": len(pool.active_proxies),
                "proxies": pool.proxies
            }
        except Exception as e:
            return {"error": str(e)}

    def refresh_proxy_pool(self):
        try:
            import asyncio
            from aegisScout.utils.proxy_pool import ProxyPoolManager
            pool = ProxyPoolManager.get_instance()
            loop = asyncio.new_event_loop()
            res = loop.run_until_complete(pool.refresh_active_pool())
            loop.close()
            return {"success": True, "results": res}
        except Exception as e:
            return {"error": str(e)}

    def is_configured(self) -> dict:
        """
        Return ONLY booleans and safe non-sensitive configuration.
        NEVER returns API keys, passwords, or tokens. Safe to call from JS.
        """
        from aegisScout.core.toml_config import config_data
        try:
            mod_b_ack = bool(
                config_data.get("outreach", {}).get("mod_b_acknowledged", False)
            )
        except Exception:
            mod_b_ack = False

        # Booleans only — no key material.
        configured = {
            "deepseek": bool(settings.deepseek_api_key),
            "openai": bool(settings.openai_api_key),
            "anthropic": bool(settings.anthropic_api_key),
            "openrouter": bool(settings.openrouter_api_key),
            "gemini": bool(settings.gemini_api_key),
            "groq": bool(settings.groq_api_key),
            "mistral": bool(settings.mistral_api_key),
            "ollama": bool(settings.ollama_base_url and settings.ollama_base_url != ""),
            "google_places": bool(settings.google_places_api_key),
            "google_custom_search": bool(
                settings.google_custom_search_api_key and settings.google_custom_search_cx
            ),
            "telegram": bool(
                settings.telegram_bot_token and settings.telegram_chat_id
            ),
            "email": bool(
                settings.notify_email_smtp_host
                and settings.notify_email_username
                and settings.notify_email_password
            ),
            "imap": bool(
                settings.notify_email_imap_host
                and settings.notify_email_imap_username
                and settings.notify_email_imap_password
            ),
            "scraper": bool(settings.scraper_api_key),
            "zenrows": bool(settings.zenrows_api_key),
            "crawlbase": bool(settings.crawlbase_api_key),
            "apify": bool(settings.apify_api_key),
            "linkedin": bool(settings.linkedin_session_cookie),
            "instagram_username": bool(settings.instagram_username),
            "instagram_password": bool(
                settings.instagram_password or self._ig_session_password
            ),
            "instagram_session_encryption_key": bool(
                settings.instagram_session_encryption_key
            ),
            "ig_session_loaded": Path("data/sessions/session.json").exists(),
        }
        return {
            "configured": configured,
            "llm_primary_provider": settings.llm_primary_provider,
            "llm_fallback_provider": settings.llm_fallback_provider or "",
            "discovery_primary_provider": settings.discovery_primary_provider,
            "outreach_mode": settings.outreach_mode,
            "max_daily_outreach": settings.max_daily_outreach,
            "ollama_base_url": settings.ollama_base_url or "http://localhost:11434",
            "ollama_model": settings.ollama_model or "llama3.2:3b",
            "notify_email_smtp_port": settings.notify_email_smtp_port or 587,
            "notify_email_imap_port": settings.notify_email_imap_port or 993,
            "proxy_pool": settings.proxy_pool or "",
            "mod_b_acknowledged": mod_b_ack,
            "automation_authorized": Path(
                "data/sessions/automation_authorized.flag"
            ).exists(),
            "ig_session_active": self._ig_session_password is not None
            or Path("data/sessions/session.json").exists(),
            "show_console": bool(config_data.get("app", {}).get("show_console", True)),
        }

    def get_settings(self):
        return self.is_configured()

    def open_logs_folder(self) -> dict:
        """Open the logs folder in Windows Explorer / system default file manager."""
        try:
            from aegisScout.utils.logger import LOGS_DIR
            import os
            import sys
            
            if not LOGS_DIR.exists():
                LOGS_DIR.mkdir(parents=True, exist_ok=True)
                
            if sys.platform == "win32":
                os.startfile(str(LOGS_DIR))
            elif sys.platform == "darwin":
                import subprocess
                subprocess.run(["open", str(LOGS_DIR)])
            else: # linux
                import subprocess
                subprocess.run(["xdg-open", str(LOGS_DIR)])
                
            return {"success": True}
        except Exception as e:
            logger.error(f"Failed to open logs folder: {e}")
            return {"error": str(e)}

    def set_config_value(self, key: str, value) -> dict:
        """
        Persist a single config key. Sensitive keys are written to .env,
        non-sensitive ones to config.toml. Refuses unknown keys.
        """
        from aegisScout.core.toml_config import config_data
        import aegisScout.core.config as _cfg_module
        from aegisScout.core.config import get_env_path

        if not isinstance(key, str) or not key.strip():
            return {"error": "key must be a non-empty string"}

        # Sensitive key — write to .env via Settings singleton.
        if key in _ENV_KEYS:
            try:
                env_path = Path(get_env_path())
                if env_path.exists() and env_path.is_dir():
                    return {"error": f".env is a directory at {env_path.absolute()}"}

                existing_lines: list[str] = []
                if env_path.exists() and env_path.is_file():
                    with open(env_path, "r", encoding="utf-8") as f:
                        existing_lines = f.readlines()

                key_positions: dict[str, int] = {}
                for i, line in enumerate(existing_lines):
                    stripped = line.strip()
                    if stripped and not stripped.startswith("#") and "=" in stripped:
                        k = stripped.split("=", 1)[0].strip().upper()
                        key_positions[k] = i

                upper_key = key.upper()
                write_val = str(value) if value is not None else ""
                new_line = f"{upper_key}={write_val}\n"
                updated = list(existing_lines)
                if upper_key in key_positions:
                    updated[key_positions[upper_key]] = new_line
                else:
                    if updated and not updated[-1].endswith("\n"):
                        updated.append("\n")
                    updated.append(new_line)

                with open(env_path, "w", encoding="utf-8") as f:
                    f.writelines(updated)

                # Reload Settings singleton so future reads see the new value.
                new_cfg = _cfg_module.Settings()
                if key in new_cfg.model_fields:
                    setattr(_cfg_module.settings, key, getattr(new_cfg, key))
                return {"success": True, "target": ".env", "key": key}
            except Exception as e:
                logger.error(f"set_config_value({key}) failed: {e}")
                return {"error": str(e)}

        # Non-sensitive key — write to config.toml.
        if key in _TOML_KEYS:
            section, toml_key = _TOML_KEYS[key]
            try:
                # Use the toml_config module's CONFIG_FILE so test fixtures
                # (which monkey-patch it) are respected and we never write
                # outside the user's real config dir.
                from aegisScout.core import toml_config as _toml_cfg
                cfg_path = Path(_toml_cfg.CONFIG_FILE)
                if cfg_path.exists():
                    data = _toml_cfg.load_toml_config(cfg_path)
                else:
                    data = {}
                if section not in data or not isinstance(data[section], dict):
                    data[section] = {}
                # Coerce types
                if toml_key in {
                    "instagram_actions_per_hour",
                    "instagram_actions_per_day",
                    "scrape_delay_min_seconds",
                    "scrape_delay_max_seconds",
                    "default_search_radius_km",
                }:
                    data[section][toml_key] = int(value) if value not in (None, "") else 0
                elif toml_key == "mod_b_acknowledged":
                    data[section][toml_key] = bool(value)
                else:
                    data[section][toml_key] = value

                # Persist to disk
                cfg_path.parent.mkdir(parents=True, exist_ok=True)
                _toml_cfg.dump_toml(data, cfg_path)
                # Update in-memory cache
                if section not in config_data or not isinstance(config_data[section], dict):
                    config_data[section] = {}
                config_data[section][toml_key] = data[section][toml_key]
                # Best-effort: keep the toml_config module's cache in sync so
                # subsequent reads via that module see the new value.
                if hasattr(_toml_cfg, "config_data") and isinstance(_toml_cfg.config_data, dict):
                    if section not in _toml_cfg.config_data or not isinstance(_toml_cfg.config_data[section], dict):
                        _toml_cfg.config_data[section] = {}
                    _toml_cfg.config_data[section][toml_key] = data[section][toml_key]
                if key == "app.show_console":
                    set_console_visibility(bool(value))
                return {"success": True, "target": "config.toml", "key": key}
            except Exception as e:
                logger.error(f"set_config_value({key}) failed: {e}")
                return {"error": str(e)}

        return {
            "error": f"Unknown config key: '{key}'. Sensitive keys (passwords/tokens/api_keys) "
                     "must be set via login_instagram() or by editing .env directly."
        }

    def save_settings(self, new_settings) -> dict:
        """
        DEPRECATED bulk setter. Kept for backward compat. Refuses any sensitive
        key. Use `set_config_value()` per key instead, or
        `login_instagram()` for credentials.
        """
        if not isinstance(new_settings, dict):
            return {"error": "new_settings must be a dict"}
        refused = []
        results = []
        for key, val in new_settings.items():
            # Block any key that could carry a secret
            lk = key.lower()
            if any(
                needle in lk
                for needle in ("api_key", "password", "token", "secret", "encryption_key")
            ):
                refused.append(key)
                continue
            if key in _ENV_KEYS:
                refused.append(key)
                continue
            res = self.set_config_value(key, val)
            results.append({"key": key, **res})
        if refused:
            return {
                "success": False,
                "error": (
                    f"Refused sensitive keys: {refused}. Use login_instagram() or "
                    "edit .env directly for secrets."
                ),
                "applied": results,
            }
        return {"success": True, "applied": results}

    # -------------------------------------------------------------------
    # UI state persistence (config.toml [app] section)
    # -------------------------------------------------------------------
    # Whitelist of UI state keys we allow JS to persist. Anything outside
    # this map is silently dropped — UI state must never leak into
    # configuration semantics.
    _UI_STATE_KEYS = {
        "theme": "last_theme",
        "language": "last_language",
        "sidebar": "last_sidebar",
        "last_view": "last_view",
        "active_tab": "last_active_tab",
        "search_sector": "last_search_sector",
        "search_location": "last_search_location",
        "search_radius": "last_search_radius",
        "search_provider": "last_search_provider",
        "leads_filter_status": "last_leads_filter_status",
        "leads_filter_has_website": "last_leads_filter_has_website",
        "leads_filter_has_instagram": "last_leads_filter_has_instagram",
    }

    def save_ui_state(self, state_json) -> dict:
        """
        Persist a UI state snapshot to config.toml under [app].

        ``state_json`` is a JSON string like
        ``{"theme": "neon", "language": "tr", "sidebar": "leads",
           "last_view": "grid"}``.
        Unknown keys are dropped. Returns the keys actually written.
        """
        from aegisScout.core import toml_config as _toml_cfg
        from aegisScout.core.toml_config import config_data

        if not isinstance(state_json, str) or not state_json.strip():
            return {"error": "state_json must be a non-empty JSON string"}

        try:
            payload = json.loads(state_json)
        except (ValueError, TypeError) as e:
            return {"error": f"Invalid JSON: {e}"}

        if not isinstance(payload, dict):
            return {"error": "state_json must decode to a JSON object"}

        written: list[str] = []
        try:
            cfg_path = Path(_toml_cfg.CONFIG_FILE)
            data: dict = {}
            if cfg_path.exists():
                data = _toml_cfg.load_toml_config(cfg_path) or {}
            if not isinstance(data, dict):
                data = {}
            data.setdefault("app", {})
            if not isinstance(data["app"], dict):
                data["app"] = {}

            for src_key, toml_key in self._UI_STATE_KEYS.items():
                if src_key not in payload:
                    continue
                value = payload[src_key]
                if value is None:
                    data["app"].pop(toml_key, None)
                    written.append(toml_key)
                    continue
                data["app"][toml_key] = str(value)
                written.append(toml_key)

            cfg_path.parent.mkdir(parents=True, exist_ok=True)
            _toml_cfg.dump_toml(data, cfg_path)

            if hasattr(_toml_cfg, "config_data") and isinstance(_toml_cfg.config_data, dict):
                _toml_cfg.config_data.setdefault("app", {})
                if not isinstance(_toml_cfg.config_data["app"], dict):
                    _toml_cfg.config_data["app"] = {}
                for toml_key in written:
                    _toml_cfg.config_data["app"][toml_key] = data["app"].get(toml_key)

            if isinstance(config_data, dict):
                config_data.setdefault("app", {})
                if not isinstance(config_data["app"], dict):
                    config_data["app"] = {}
                for toml_key in written:
                    config_data["app"][toml_key] = data["app"].get(toml_key)

            return {"success": True, "written": written}
        except Exception as e:
            logger.error(f"save_ui_state failed: {e}")
            return {"error": str(e)}

    def load_ui_state(self) -> str:
        """
        Return the saved UI state as a JSON string.

        Reads config.toml [app] and returns only the keys listed in
        ``_UI_STATE_KEYS``. If nothing is persisted, returns ``"{}"`` so
        the JS layer always has a valid JSON object to parse.
        """
        try:
            from aegisScout.core.toml_config import load_toml_config
            section: dict = {}
            try:
                section = load_toml_config().get("app", {}) or {}
            except Exception:
                section = {}
            if not isinstance(section, dict):
                section = {}

            payload: dict[str, str] = {}
            inverse = {v: k for k, v in self._UI_STATE_KEYS.items()}
            for toml_key, src_key in inverse.items():
                if toml_key in section and section[toml_key] is not None:
                    payload[src_key] = str(section[toml_key])

            return json.dumps(payload, ensure_ascii=False)
        except Exception as e:
            logger.error(f"load_ui_state failed: {e}")
            return "{}"

    # -------------------------------------------------------------------
    # Console window toggle (JS bridge)
    # -------------------------------------------------------------------
    def toggle_console(self, visible: bool | None = None) -> bool:
        """
        Toggle the Windows console window attached to this process.

        ``visible`` is an optional bool: ``True`` to show, ``False`` to
        hide, ``None`` to flip the current visibility. Returns True only
        when an actual ShowWindow call was issued. On non-Windows
        platforms, or when the process has no console (typical for
        PyInstaller frozen builds), the function returns False silently.
        """
        if sys.platform != "win32":
            logger.info("toggle_console: non-Windows platform, skipping.")
            return False
        try:
            hwnd = ctypes.windll.kernel32.GetConsoleWindow()
        except AttributeError as e:
            logger.error(f"toggle_console: ctypes.windll unavailable: {e}")
            return False
        if not hwnd:
            logger.warning(
                "toggle_console: no console attached to this process "
                "(common in PyInstaller builds); skipping."
            )
            return False
        try:
            target_state = 5 if visible is None or bool(visible) else 0
            ctypes.windll.user32.ShowWindow(hwnd, target_state)
            return True
        except Exception as e:
            logger.error(f"toggle_console: ShowWindow failed: {e}")
            return False

    # -------------------------------------------------------------------
    # Instagram login (in-memory only)
    # -------------------------------------------------------------------
    def login_instagram(self, username: str, password: str) -> dict:
        """
        Authenticate with Instagram using credentials supplied at runtime.
        Credentials are kept in MEMORY ONLY (never persisted to .env or toml).
        Returns success status — JS receives only the boolean result.
        """
        if not username or not password:
            return {"success": False, "error": "Kullanıcı adı ve şifre gerekli."}

        self._ig_session_password = password  # memory-only, will be cleared
        try:
            try:
                from aegisScout.outreach.instagram_client import InstagramClient
            except ImportError:
                self._ig_session_password = None
                return {
                    "success": False,
                    "error": (
                        "Mod B (instagrapi) yüklü değil. "
                        "`pip install instagrapi` veya `[mod-b]` extras."
                    ),
                }

            # Build a temporary settings snapshot to inject credentials
            # without writing to disk.
            from aegisScout.core.config import settings as _settings_singleton
            original_user = _settings_singleton.instagram_username
            original_pass = _settings_singleton.instagram_password
            _settings_singleton.instagram_username = username
            _settings_singleton.instagram_password = password

            try:
                ig = InstagramClient()
                ok = ig.login()
                if ok:
                    return {"success": True, "username": username}
                return {
                    "success": False,
                    "error": "Instagram girişi başarısız. Şifre/2FA/challenge kontrolü yapın.",
                }
            finally:
                # Always restore the singleton and clear in-memory password.
                _settings_singleton.instagram_username = original_user
                _settings_singleton.instagram_password = original_pass
                self._ig_session_password = None
        except Exception as e:
            logger.error(f"login_instagram failed: {e}")
            self._ig_session_password = None
            return {"success": False, "error": str(e)}

    def logout_instagram(self) -> dict:
        """Forget the in-memory Instagram session and delete the encrypted session file."""
        self._ig_session_password = None
        try:
            session_file = Path("data/sessions/session.json")
            if session_file.exists():
                session_file.unlink()
            flag_file = Path("data/sessions/automation_authorized.flag")
            if flag_file.exists():
                flag_file.unlink()
            return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    # -------------------------------------------------------------------
    # Safe command dispatcher — READ-ONLY operations only.
    # -------------------------------------------------------------------
    def trigger_command(self, command: str, args: dict | None = None) -> dict:
        """
        Dispatch a SAFE read-only command from JS. Refuses any command that
        mutates state or sends messages.
        """
        if not isinstance(command, str) or command not in _SAFE_COMMANDS:
            return {
                "error": f"Command '{command}' is not in the safe whitelist.",
                "allowed": sorted(_SAFE_COMMANDS),
            }

        args = args or {}

        try:
            if command == "list":
                return self.get_leads(args.get("status"))
            if command == "discover":
                return self.discover_leads(
                    args.get("sector", ""),
                    args.get("location", ""),
                    args.get("radius", 10),
                    args.get("provider", settings.discovery_primary_provider),
                )
            if command == "research":
                return self.research_lead(args.get("lead_id"), force=args.get("force", True))
            if command == "export":
                return {"queued": False, "note": "Use aegisScout export from CLI; not exposed in GUI."}
            if command == "is_configured":
                return self.is_configured()
            if command == "get_stats":
                return self.get_stats()
            if command == "get_leads":
                return self.get_leads(args.get("status"))
            if command == "get_lead_details":
                return self.get_lead_details(args.get("lead_id"))
            if command == "get_campaigns":
                return self.get_campaigns()
            if command == "get_campaign_details":
                return self.get_campaign_details(args.get("campaign_id"))
            if command == "get_activity_logs":
                return self.get_activity_logs()
            if command == "get_sessions":
                return self.get_sessions()
            return {"error": f"Unhandled safe command: {command}"}
        except Exception as e:
            return {"error": str(e)}

    # -------------------------------------------------------------------
    # Multi-platform social discovery (additive — only fills empty fields)
    # -------------------------------------------------------------------
    def discover_social_profiles(self, lead_id) -> dict:
        """
        Run the YouTube / LinkedIn / TikTok / Facebook / Telegram / X-Twitter
        discovery for a single lead and persist anything we find.

        Existing URLs on the lead row are NEVER overwritten — the discovery
        is purely additive. Returns a dict with the updated lead details so
        the GUI can re-render without an extra round-trip.
        """
        try:
            lead_id = int(lead_id)
        except (TypeError, ValueError):
            return {"error": "lead_id must be an integer."}

        from aegisScout.discovery.social_discovery import SocialDiscovery

        def run_discovery():
            try:
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                try:
                    loop.run_until_complete(
                        commands.enrich_lead_socials(lead_id, SocialDiscovery())
                    )
                finally:
                    loop.close()
                if self._window:
                    try:
                        self._window.evaluate_js(
                            f"finishSocialDiscovery({lead_id}, true, null)"
                        )
                    except Exception:
                        pass
            except Exception as ex:
                logger.error(f"Social discovery failed for lead {lead_id}: {ex}")
                if self._window:
                    try:
                        err_json = json.dumps(str(ex))
                        self._window.evaluate_js(
                            f"finishSocialDiscovery({lead_id}, false, {err_json})"
                        )
                    except Exception:
                        pass

        t = threading.Thread(target=run_discovery, daemon=True)
        t.start()
        return {"queued": True}

    # -------------------------------------------------------------------
    # Campaigns (unchanged)
    # -------------------------------------------------------------------
    def get_campaigns(self):
        try:
            from sqlmodel import func

            with Session(engine) as session:
                campaigns = session.exec(
                    select(Campaign)
                    .where(Campaign.session_id == self._active_session_id)
                    .order_by(Campaign.id.desc())
                ).all()
                res = []
                for c in campaigns:
                    leads_count = session.exec(
                        select(func.count(Lead.id)).where(Lead.campaign_id == c.id)
                    ).one() or 0

                    replied_count = session.exec(
                        select(func.count(Lead.id)).where(
                            (Lead.campaign_id == c.id) & (Lead.status == "replied")
                        )
                    ).one() or 0

                    contacted_count = session.exec(
                        select(func.count(Lead.id)).where(
                            (Lead.campaign_id == c.id)
                            & (Lead.status.in_(["contacted", "replied"]))
                        )
                    ).one() or 0

                    c_data = c.model_dump()
                    c_data["created_at"] = c_data["created_at"].strftime("%Y-%m-%d %H:%M:%S")
                    c_data["total_leads"] = leads_count
                    c_data["contacted_leads"] = contacted_count
                    c_data["replied_leads"] = replied_count
                    res.append(c_data)
                return res
        except Exception as e:
            return {"error": str(e)}

    def create_campaign(self, name, sector_filter=None, location_filter=None):
        try:
            with Session(engine) as session:
                exists = session.exec(
                    select(Campaign).where(
                        (Campaign.name == name)
                        & (Campaign.session_id == self._active_session_id)
                    )
                ).first()
                if exists:
                    return {"error": f"'{name}' adında bir kampanya zaten mevcut."}

                campaign = Campaign(
                    name=name,
                    sector_filter=sector_filter or None,
                    location_filter=location_filter or None,
                    session_id=self._active_session_id,
                )
                session.add(campaign)

                log = ActivityLog(
                    action="campaign_create",
                    details=f"Kampanya oluşturuldu (GUI): {name}",
                    session_id=self._active_session_id,
                )
                session.add(log)
                session.commit()
                return {"success": True, "id": campaign.id}
        except Exception as e:
            return {"error": str(e)}

    def assign_lead_to_campaign(self, campaign_id, lead_id=None, auto_filter=False):
        try:
            with Session(engine) as session:
                campaign = session.get(Campaign, campaign_id)
                if not campaign:
                    return {"error": "Kampanya bulunamadı."}

                assigned_count = 0
                if lead_id:
                    lead = session.get(Lead, lead_id)
                    if not lead:
                        return {"error": "Lead bulunamadı."}
                    lead.campaign_id = campaign.id
                    session.add(lead)
                    assigned_count = 1

                if auto_filter:
                    stmt = select(Lead).where(
                        (Lead.campaign_id.is_(None))
                        & (Lead.session_id == self._active_session_id)
                    )
                    if campaign.sector_filter:
                        stmt = stmt.where(Lead.sector == campaign.sector_filter)
                    if campaign.location_filter:
                        stmt = stmt.where(Lead.address.like(f"%{campaign.location_filter}%"))

                    leads_to_assign = session.exec(stmt).all()
                    for lead in leads_to_assign:
                        lead.campaign_id = campaign.id
                        session.add(lead)
                        assigned_count += 1

                if assigned_count > 0:
                    log = ActivityLog(
                        action="campaign_assign",
                        details=f"{assigned_count} adet lead Kampanya '{campaign.name}' (ID: {campaign.id}) atandı (GUI).",
                        session_id=self._active_session_id,
                    )
                    session.add(log)
                    session.commit()
                    return {"success": True, "assigned": assigned_count}
                else:
                    return {"success": True, "assigned": 0, "message": "Atanacak yeni aday bulunamadı."}
        except Exception as e:
            return {"error": str(e)}

    def get_campaign_details(self, campaign_id):
        try:
            with Session(engine) as session:
                campaign = session.get(Campaign, campaign_id)
                if not campaign:
                    return {"error": "Kampanya bulunamadı."}

                leads = session.exec(
                    select(Lead).where(Lead.campaign_id == campaign.id).order_by(Lead.id.desc())
                ).all()
                leads_data = [l.model_dump() for l in leads]
                for l in leads_data:
                    l["discovered_at"] = l["discovered_at"].strftime("%Y-%m-%d %H:%M:%S")
                    l["updated_at"] = l["updated_at"].strftime("%Y-%m-%d %H:%M:%S")

                campaign_data = campaign.model_dump()
                campaign_data["created_at"] = campaign_data["created_at"].strftime("%Y-%m-%d %H:%M:%S")

                return {
                    "campaign": campaign_data,
                    "leads": leads_data,
                }
        except Exception as e:
            return {"error": str(e)}

    # -------------------------------------------------------------------
    # Logs / sessions / housekeeping (unchanged)
    # -------------------------------------------------------------------
    def get_activity_logs(self):
        try:
            with Session(engine) as session:
                logs = session.exec(
                    select(ActivityLog)
                    .where(ActivityLog.session_id == self._active_session_id)
                    .order_by(ActivityLog.timestamp.desc())
                ).all()
                res = []
                for log in logs:
                    l_data = log.model_dump()
                    l_data["timestamp"] = l_data["timestamp"].strftime("%Y-%m-%d %H:%M:%S")
                    res.append(l_data)
                return res
        except Exception as e:
            return {"error": str(e)}

    def get_stats(self):
        try:
            from sqlmodel import func

            with Session(engine) as session:
                statuses = ["new", "researched", "contacted", "replied", "rejected", "do_not_contact"]
                counts = {}
                for s in statuses:
                    cnt = session.exec(
                        select(func.count(Lead.id)).where(
                            (Lead.session_id == self._active_session_id)
                            & (Lead.status == s)
                        )
                    ).one()
                    counts[s] = cnt or 0
                total = session.exec(
                    select(func.count(Lead.id)).where(Lead.session_id == self._active_session_id)
                ).one() or 0
                counts["total"] = total
                counts["drafted"] = session.exec(
                    select(func.count(Lead.id)).where(
                        (Lead.session_id == self._active_session_id)
                        & (Lead.status == "drafted")
                    )
                ).one() or 0
                return counts
        except Exception as e:
            return {"error": str(e)}

    def clear_activity_logs(self):
        try:
            with Session(engine) as session:
                session.execute(
                    text("DELETE FROM activity_log WHERE session_id = :sid"),
                    {"sid": self._active_session_id},
                )
                log = ActivityLog(
                    action="session_start",
                    details="Yeni oturum başlatıldı ve geçmiş temizlendi.",
                    session_id=self._active_session_id,
                )
                session.add(log)
                session.commit()
                return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    def create_activity_log(self, action, details):
        try:
            with Session(engine) as session:
                log = ActivityLog(action=action, details=details, session_id=self._active_session_id)
                session.add(log)
                session.commit()
                return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    def get_sessions(self):
        try:
            with Session(engine) as session:
                sessions = session.exec(select(UserSession).order_by(UserSession.id.desc())).all()
                res = []
                for s in sessions:
                    s_data = s.model_dump()
                    s_data["created_at"] = s_data["created_at"].strftime("%Y-%m-%d %H:%M:%S")
                    s_data["is_active"] = (s.id == self._active_session_id)
                    res.append(s_data)
                return res
        except Exception as e:
            return {"error": str(e)}

    def create_session(self, name):
        try:
            if not name or not name.strip():
                return {"error": "Oturum adı boş bırakılamaz."}
            with Session(engine) as session:
                new_s = UserSession(name=name.strip())
                session.add(new_s)
                session.commit()
                self._active_session_id = new_s.id
                return {"success": True, "session_id": new_s.id}
        except Exception as e:
            return {"error": str(e)}

    def switch_session(self, session_id):
        try:
            session_id = int(session_id)
            with Session(engine) as session:
                s = session.get(UserSession, session_id)
                if not s:
                    return {"error": f"Oturum ID {session_id} bulunamadı."}
                self._active_session_id = s.id
                return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    def delete_session(self, session_id):
        try:
            session_id = int(session_id)
            if session_id == 1:
                return {"error": "Varsayılan çalışma oturumu silinemez."}
            if session_id == self._active_session_id:
                return {"error": "Aktif çalışma oturumu silinemez. Lütfen önce başka bir oturuma geçiş yapın."}

            with Session(engine) as session:
                s = session.get(UserSession, session_id)
                if not s:
                    return {"error": "Oturum bulunamadı."}
                session.execute(text("DELETE FROM leads WHERE session_id = :sid"), {"sid": session_id})
                session.execute(text("DELETE FROM campaigns WHERE session_id = :sid"), {"sid": session_id})
                session.execute(text("DELETE FROM activity_log WHERE session_id = :sid"), {"sid": session_id})
                session.delete(s)
                session.commit()
                return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    def rename_session(self, session_id, new_name):
        try:
            session_id = int(session_id)
            new_name = (new_name or "").strip()
            if not new_name:
                return {"error": "Oturum adı boş bırakılamaz."}
            if len(new_name) > 80:
                return {"error": "Oturum adı en fazla 80 karakter olabilir."}
            with Session(engine) as session:
                s = session.get(UserSession, session_id)
                if not s:
                    return {"error": f"Oturum ID {session_id} bulunamadı."}
                s.name = new_name
                session.add(s)
                log = ActivityLog(
                    action="session_rename",
                    details=f"Oturum yeniden adlandırıldı → '{new_name}'",
                    session_id=self._active_session_id,
                )
                session.add(log)
                session.commit()
                return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    def delete_lead(self, lead_id):
        try:
            lead_id = int(lead_id)
            with Session(engine) as session:
                lead = session.get(Lead, lead_id)
                if not lead:
                    return {"error": "Aday bulunamadı."}

                session.execute(text("DELETE FROM research_notes WHERE lead_id = :lid"), {"lid": lead_id})
                session.execute(text("DELETE FROM messages WHERE lead_id = :lid"), {"lid": lead_id})
                session.delete(lead)

                log = ActivityLog(
                    action="lead_delete",
                    details=f"Aday silindi: {lead.business_name}",
                    session_id=self._active_session_id,
                )
                session.add(log)
                session.commit()
                return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    def delete_campaign(self, campaign_id):
        try:
            campaign_id = int(campaign_id)
            with Session(engine) as session:
                campaign = session.get(Campaign, campaign_id)
                if not campaign:
                    return {"error": "Kampanya bulunamadı."}
                session.execute(
                    text("UPDATE leads SET campaign_id = NULL WHERE campaign_id = :cid"),
                    {"cid": campaign_id},
                )
                session.delete(campaign)

                log = ActivityLog(
                    action="campaign_delete",
                    details=f"Kampanya silindi: {campaign.name}",
                    session_id=self._active_session_id,
                )
                session.add(log)
                session.commit()
                return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    def delete_activity_log(self, log_id):
        try:
            log_id = int(log_id)
            with Session(engine) as session:
                log = session.get(ActivityLog, log_id)
                if not log:
                    return {"error": "Günlük kaydı bulunamadı."}
                session.delete(log)
                session.commit()
                return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    # -------------------------------------------------------------------
    # B2B Growth Platform V2 Features Bridge
    # -------------------------------------------------------------------
    def verify_email_address(self, email):
        try:
            from aegisScout.utils.email_verifier import verify_email
            return verify_email(email)
        except Exception as e:
            return {"error": str(e)}

    def run_screen_audit(self, lead_id):
        try:
            from aegisScout.core.screen_audit import run_website_screen_audit
            import asyncio
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            res = loop.run_until_complete(run_website_screen_audit(int(lead_id)))
            loop.close()
            return res
        except Exception as e:
            return {"error": str(e)}

    def run_waterfall_lead(self, lead_id):
        try:
            from aegisScout.core.waterfall import run_waterfall_enrichment
            import asyncio
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            res = loop.run_until_complete(run_waterfall_enrichment(int(lead_id)))
            loop.close()
            return res
        except Exception as e:
            return {"error": str(e)}

    def get_waterfall_config(self):
        try:
            from aegisScout.core.waterfall import load_waterfall_config
            return load_waterfall_config()
        except Exception as e:
            return {"error": str(e)}

    def save_waterfall_config(self, config_list):
        try:
            from aegisScout.core.waterfall import save_waterfall_config
            save_waterfall_config(config_list)
            return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    def get_warmup_status(self):
        try:
            from aegisScout.outreach.warmup import _load_stats
            return {
                "active": settings.email_warmup_active,
                "stats": _load_stats()
            }
        except Exception as e:
            return {"error": str(e)}

    def toggle_warmup(self, is_active):
        try:
            from aegisScout.core.config import set_settings, get_settings
            set_settings(email_warmup_active=bool(is_active))
            return {"success": True, "active": get_settings().email_warmup_active}
        except Exception as e:
            return {"error": str(e)}

    def run_warmup_cycle_manual(self):
        try:
            from aegisScout.outreach.warmup import run_p2p_warmup_cycle
            import asyncio
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            res = loop.run_until_complete(run_p2p_warmup_cycle())
            loop.close()
            return res
        except Exception as e:
            return {"error": str(e)}

    def launch_whatsapp_auto(self, lead_id, message_text):
        try:
            with Session(engine) as session:
                lead = session.get(Lead, int(lead_id))
                if not lead or not lead.phone:
                    return {"error": "Aday veya telefon numarası bulunamadı."}
                from aegisScout.outreach.browser_automation import send_whatsapp_message_auto
                res = send_whatsapp_message_auto(lead.phone, message_text)
                if res.get("success"):
                    new_msg = Message(
                        lead_id=lead.id,
                        direction="outbound",
                        channel="whatsapp_auto",
                        content=message_text,
                        status="sent",
                        sent_at=_utcnow()
                    )
                    lead.status = "contacted"
                    lead.updated_at = _utcnow()
                    session.add(new_msg)
                    session.add(lead)
                    session.commit()
                return res
        except Exception as e:
            return {"error": str(e)}

    def launch_whatsapp_login(self):
        try:
            from aegisScout.outreach.browser_automation import launch_whatsapp_session
            launch_whatsapp_session()
            return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    def launch_linkedin_login(self):
        try:
            from aegisScout.outreach.browser_automation import launch_linkedin_session
            launch_linkedin_session()
            return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    def launch_linkedin_auto(self, lead_id, message_text):
        try:
            with Session(engine) as session:
                lead = session.get(Lead, int(lead_id))
                if not lead or not lead.linkedin_url:
                    return {"error": "Aday veya LinkedIn adresi bulunamadı."}
                from aegisScout.outreach.browser_automation import send_linkedin_connection_auto
                res = send_linkedin_connection_auto(lead.linkedin_url, message_text)
                if res.get("success"):
                    new_msg = Message(
                       lead_id=lead.id,
                       direction="outbound",
                       channel="linkedin_auto",
                       content=message_text,
                       status="sent",
                       sent_at=_utcnow()
                    )
                    lead.status = "contacted"
                    lead.updated_at = _utcnow()
                    session.add(new_msg)
                    session.add(lead)
                    session.commit()
                return res
        except Exception as e:
            return {"error": str(e)}

    def get_unibox_messages(self):
        try:
            with Session(engine) as session:
                stmt = select(Message).order_by(Message.created_at.asc())
                messages = session.exec(stmt).all()
                grouped = {}
                for m in messages:
                    lead = session.get(Lead, m.lead_id)
                    lead_name = lead.business_name if lead else "Bilinmeyen Aday"
                    lead_email = lead.email if lead else None
                    lead_phone = lead.phone if lead else None
                    
                    if m.lead_id not in grouped:
                        grouped[m.lead_id] = {
                            "lead_id": m.lead_id,
                            "lead_name": lead_name,
                            "lead_email": lead_email,
                            "lead_phone": lead_phone,
                            "messages": []
                        }
                    
                    md = m.model_dump()
                    if md.get("sent_at"):
                        md["sent_at"] = md["sent_at"].strftime("%Y-%m-%d %H:%M:%S")
                    if md.get("created_at"):
                        md["created_at"] = md["created_at"].strftime("%Y-%m-%d %H:%M:%S")
                        
                    grouped[m.lead_id]["messages"].append(md)
                    
                return list(grouped.values())
        except Exception as e:
            return {"error": str(e)}

    def send_unibox_reply(self, lead_id, channel, content):
        try:
            lead_id = int(lead_id)
            with Session(engine) as session:
                lead = session.get(Lead, lead_id)
                if not lead:
                    return {"error": "Aday bulunamadı."}
                
                if channel == "email":
                    if not lead.email:
                        return {"error": "Adayın e-posta adresi bulunamadı."}
                    from aegisScout.outreach.email_client import send_cold_email
                    success, msg, used_smtp_id = send_cold_email(lead.email, "YNT: İletişim", content)
                    if not success:
                        return {"error": msg}
                    new_msg = Message(
                        lead_id=lead.id,
                        direction="outbound",
                        channel="email",
                        content=content,
                        status="sent",
                        sent_at=_utcnow(),
                        smtp_account_id=used_smtp_id
                    )
                    session.add(new_msg)
                    session.commit()
                    return {"success": True}
                    
                elif channel == "whatsapp":
                    if not lead.phone:
                        return {"error": "Adayın telefon numarası bulunamadı."}
                    from aegisScout.outreach.browser_automation import send_whatsapp_message_auto
                    res = send_whatsapp_message_auto(lead.phone, content)
                    if res.get("success"):
                        new_msg = Message(
                            lead_id=lead.id,
                            direction="outbound",
                            channel="whatsapp_auto",
                            content=content,
                            status="sent",
                            sent_at=_utcnow()
                        )
                        session.add(new_msg)
                        session.commit()
                    return res
                    
                elif channel == "linkedin":
                    if not lead.linkedin_url:
                        return {"error": "Adayın LinkedIn adresi bulunamadı."}
                    from aegisScout.outreach.browser_automation import send_linkedin_connection_auto
                    res = send_linkedin_connection_auto(lead.linkedin_url, content)
                    if res.get("success"):
                        new_msg = Message(
                            lead_id=lead.id,
                            direction="outbound",
                            channel="linkedin_auto",
                            content=content,
                            status="sent",
                            sent_at=_utcnow()
                        )
                        session.add(new_msg)
                        session.commit()
                    return res
                else:
                    return {"error": f"Geçersiz kanal: {channel}"}
        except Exception as e:
            return {"error": str(e)}

    def update_lead_status(self, lead_id, new_status):
        try:
            lead_id = int(lead_id)
            valid_statuses = {
                "new", "researched", "drafted", "contacted",
                "replied", "converted", "rejected", "do_not_contact",
                "meeting_scheduled"
            }
            if new_status not in valid_statuses:
                return {"error": f"Geçersiz durum: '{new_status}'."}
            with Session(engine) as session:
                lead = session.get(Lead, lead_id)
                if not lead:
                    return {"error": "Aday bulunamadı."}
                lead.status = new_status
                lead.updated_at = _utcnow()
                session.add(lead)

                log = ActivityLog(
                    action="lead_status_update",
                    details=f"Aday durumu güncellendi: {lead.business_name} -> {new_status}",
                    session_id=self._active_session_id,
                )
                session.add(log)
                session.commit()
                return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    def clear_all_leads(self):
        try:
            with Session(engine) as session:
                session.execute(
                    text("DELETE FROM research_notes WHERE lead_id IN (SELECT id FROM leads WHERE session_id = :sid)"),
                    {"sid": self._active_session_id},
                )
                session.execute(
                    text("DELETE FROM messages WHERE lead_id IN (SELECT id FROM leads WHERE session_id = :sid)"),
                    {"sid": self._active_session_id},
                )
                session.execute(
                    text("DELETE FROM leads WHERE session_id = :sid"),
                    {"sid": self._active_session_id},
                )

                log = ActivityLog(
                    action="leads_clear_all",
                    details="Tüm adaylar temizlendi.",
                    session_id=self._active_session_id,
                )
                session.add(log)
                session.commit()
                return {"success": True}
        except Exception as e:
            return {"error": str(e)}

    def clear_leads_by_filter(self, status=None, sector=None):
        try:
            with Session(engine) as session:
                subquery_conds = ["session_id = :sid"]
                params = {"sid": self._active_session_id}

                if status and status != "all":
                    subquery_conds.append("status = :status")
                    params["status"] = status
                if sector and sector.strip() != "":
                    subquery_conds.append("sector = :sector")
                    params["sector"] = sector.strip()

                cond_str = " AND ".join(subquery_conds)

                session.execute(
                    text(f"DELETE FROM research_notes WHERE lead_id IN (SELECT id FROM leads WHERE {cond_str})"),
                    params,
                )
                session.execute(
                    text(f"DELETE FROM messages WHERE lead_id IN (SELECT id FROM leads WHERE {cond_str})"),
                    params,
                )

                result = session.execute(
                    text(f"DELETE FROM leads WHERE {cond_str}"),
                    params,
                )
                deleted_count = result.rowcount

                log = ActivityLog(
                    action="leads_clear_filter",
                    details=f"Filtreli adaylar temizlendi (Durum: {status or 'Tümü'}, Sektör: {sector or 'Tümü'}, Silinen: {deleted_count}).",
                    session_id=self._active_session_id,
                )
                session.add(log)
                session.commit()
                return {"success": True, "deleted_count": deleted_count}
        except Exception as e:
            return {"error": str(e)}

    def log_js_error(self, message, source, lineno, colno, error):
        try:
            logger.error(
                f"JS Error: {message} at {source}:{lineno}:{colno}. Error object: {error}"
            )
            log_dir = Path("logs")
            log_dir.mkdir(exist_ok=True)
            with open(log_dir / "gui_errors.txt", "a", encoding="utf-8") as f:
                f.write(
                    f"JS Error: {message} at {source}:{lineno}:{colno}. Stack: {error}\n"
                )
            return {"success": True}
        except Exception as e:
            return {"error": str(e)}


def set_console_visibility(visible: bool):
    if sys.platform != "win32":
        return
    try:
        hwnd = ctypes.windll.kernel32.GetConsoleWindow()
        logger.info(f"set_console_visibility: visible={visible}, hwnd={hwnd}")
        if hwnd:
            ctypes.windll.user32.ShowWindow(hwnd, 5 if visible else 0)
            logger.info(f"ShowWindow called with state={5 if visible else 0}")
        else:
            logger.warning("GetConsoleWindow returned NULL (no console associated with this process).")
    except Exception as e:
        logger.error(f"Failed to set console visibility to {visible}: {e}")


def close_console_window():
    if sys.platform != "win32":
        return
    try:
        hwnd = ctypes.windll.kernel32.GetConsoleWindow()
        if hwnd:
            ctypes.windll.user32.ShowWindow(hwnd, 0)
    except Exception:
        pass


def _patch_pywebview_js_path() -> None:
    """Fix pywebview JS file path resolution for frozen (PyInstaller) EXE.

    In frozen EXE, get_js_dir() fails because __file__ points inside the PYZ
    archive while data files (api.js, finish.js, etc.) are extracted to
    sys._MEIPASS. We patch the function to check sys._MEIPASS first.
    """
    if not getattr(sys, "frozen", False):
        return  # Not frozen — original logic works fine

    import webview.util as _wv_util
    import webview
    _original_get_js_dir = _wv_util.get_js_dir

    def _patched_get_js_dir() -> str:
        # In PyInstaller frozen builds, the webview module may be in PYZ
        # but data files (webview/js/) are in sys._MEIPASS.
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            # Try _MEIPASS/webview/js/ first
            candidate = os.path.join(meipass, "webview", "js")
            if os.path.isdir(candidate):
                return candidate
        # Fall back to original logic
        return _original_get_js_dir()

    _wv_util.get_js_dir = _patched_get_js_dir


def _get_icon_path() -> str | None:
    """Return path to the .ico file for the GUI window, or None if not found."""
    if getattr(sys, "frozen", False):
        return None  # PyInstaller handles icon via .spec file
    for candidate in [
        Path(__file__).parent / "assets" / "logo.ico",
        Path(__file__).parent / "assets" / "logo.png",
    ]:
        if candidate.exists():
            return str(candidate)
    return None


def _start_window(api, icon_path, debug_mode):
    """Create a single window and start the GUI loop."""
    window = webview.create_window(
        title="aegisScout — İşletme Keşif ve Satış Otomasyonu",
        html=HTML_CONTENT,
        js_api=api,
        width=1200,
        height=800,
        resizable=True,
        min_size=(1024, 700),
    )
    api._window = window
    # Auto-detect GUI (edgechromium on Windows with WebView2, mshtml fallback)
    webview.start(debug=debug_mode, private_mode=False, icon=icon_path)


def start_gui():
    # Apply PyInstaller frozen fix for pywebview JS bridge path
    _patch_pywebview_js_path()

    from aegisScout.core.toml_config import config_data
    try:
        show = bool(config_data.get("app", {}).get("show_console", True))
    except Exception:
        show = True
    set_console_visibility(show)
    icon_path = _get_icon_path()
    is_frozen = getattr(sys, "frozen", False)
    debug_mode = False if is_frozen else show

    try:
        _start_window(GuiApi(), icon_path, debug_mode)
    except Exception as e:
        logger.exception(f"GUI startup failed (attempt 1): {e}")
        # Fallback attempt with debug=False
        try:
            logger.info("Starting GUI fallback attempt...")
            _start_window(GuiApi(), icon_path, False)
        except Exception as e2:
            logger.exception(f"GUI fallback also failed: {e2}")
            print(f"[ERROR] GUI could not start: {e2}")
            print("[INFO] Use CLI mode instead: aegisScout --help")

    logger.info("GUI window closed. Restoring console visibility.")
    set_console_visibility(True)
    close_console_window()


if __name__ == "__main__":
    start_gui()
